from xlsx2csv import *
import openpyxl
from openpyxl import Workbook
#creat new .xlsx file and give the first line the titles of the columns
newElection = Workbook()
newHashtag = Workbook()
#sheet_new = newElection.active
sheet_new = newElection.create_sheet("1.Sheet",0)
sheet_hashtag = newElection.create_sheet("2.Sheet",1)
sheet_new ['A1'] = "indice"
sheet_new ['B1'] ="handle"
sheet_new ['C1'] = "text"
sheet_new ['D1'] = "time"
sheet_new ['E1'] = "retweet_count"
sheet_new ['F1'] ="fav_count"
sheet_hashtag ['A1'] = "1.Hashtag"
sheet_hashtag ['B1'] = "2.Hashtag"
sheet_hashtag ['C1'] = "3.Hashtag"
sheet_hashtag ['D1'] = "4.Hashtag"
sheet_hashtag ['E1'] = "5.Hashtag"
sheet_hashtag ['F1'] = "6.Hashtag"
oldElection = openpyxl.load_workbook('american-election-tweets.xlsx')
sheet_old = oldElection.active

#quicksort
def quicksort(myList, start, end):
    if start < end:
        # partition the list
        pivot = partition(myList, start, end)
        # sort both halves
        quicksort(myList, start, pivot-1)
        quicksort(myList, pivot+1, end)
    return myList


def partition(myList, start, end):
    pivot = myList[start]
    left = start+1
    right = end
    done = False
    while not done:
        while left <= right and myList[left] <= pivot:
            left = left + 1
        while myList[right] >= pivot and right >=left:
            right = right -1
        if right < left:
            done= True
        else:
            # swap places
            temp=myList[left]
            myList[left]=myList[right]
            myList[right]=temp
    # swap start with myList[right]
    temp=myList[start]
    myList[start]=myList[right]
    myList[right]=temp
    return right

#open the given file
oldElection = openpyxl.load_workbook('american-election-tweets.xlsx')
sheet_old = oldElection.active

	


#check if there are any Hashtags in the tweet
def hashtag_occ(text,pos):
	if (text.find("#",pos)>=0):
		if (ord(text[text.find("#",pos)])== ord("#")):	
			return text.find("#")
		else:
			hashtag_occ(text,(text.find("#",pos)))
	else:
		return -1




#collect all the Hashtags of a tweet in a list
def hashtag_names(text,pos,name,counter):	
	new_pos = text.find("#",pos)
	ending_space= text.find(" ",new_pos)
	
	ending_newline = text.find("\n",new_pos)
	
	if (ending_space == -1):
		ending_space = 1000
	if (ending_newline ==-1):
		ending_newline =1000
	if (ending_newline==ending_space==1000):
		ending = len(text) + 1
	else:
		ending = min(ending_space, ending_newline)
	temp_name= text[new_pos:ending]
	if (ord(text[new_pos])== ord("#")):
		sheet_hashtag.cell(row = j,column=(counter)).value = temp_name
	if (text.find("#",new_pos+1)) >0:
		counter = counter +1
		hashtag_names(text,text.find("#",new_pos+1),name,counter)




# parse the values of each row (with hashtag in the tweet) to the new File
j= 2
for i in range(2,6127):
#Filter Hashtags => kill all lines without a # in the tweet
	if (hashtag_occ(sheet_old.cell(row = i,column = 2).value, 0) != -1):
#Filter truncated => kill all the lines with a truncated tweet
		if (sheet_old.cell(row=i, column = 11).value == "False"):


#kill all uninteresting columns	is_retweet, is_reply, url, original_author, is_quote
				
			sheet_new.cell(row = j,column=2).value = sheet_old.cell(row=i, column=1).value
			sheet_new.cell(row = j,column=3).value = sheet_old.cell(row=i, column=2).value
			sheet_new.cell(row = j,column=4).value = sheet_old.cell(row=i, column=5).value		
			sheet_new.cell(row = j,column=5).value = sheet_old.cell(row=i, column=8).value
			sheet_new.cell(row = j,column=6).value = sheet_old.cell(row=i, column=9).value
			
			hashtag_names(sheet_old.cell(row = i,column = 2).value, 0,[],1)
			sheet_new.cell(row=j, column= 1).value = j-1	
			j=j+1
newElection.save("Election.xlsx")
savedElection = openpyxl.load_workbook('Election.xlsx')
sheet_saved = savedElection["2.Sheet"]

sheet_sorthashtag = newHashtag.create_sheet("1.Sheet",0)
sheet_sorthashtag ['A1'] = "hashtag"
sheet_sorthashtag ['B1'] = "tweet_indices"
#sheet_sorthashtag ['C1'] = "retweet_count_aufsummiert"
#sheet_sorthashtag ['D1'] = "tweet_fav_count_aufsummiert"

#get a collection of sorted hashtags (1 occurence for each one)
hashtagnames=[]
for s in range (2,1257):
	for t in range (1,6):		
		if (sheet_hashtag.cell(row=s , column = t).value != None):	
			hashtagnames.append(sheet_hashtag.cell(row=s , column = t).value)
hash_sort= quicksort(hashtagnames,0,(len(hashtagnames)-1))
		
v = 0
hash_coll =[]
hash_coll.append(hash_sort[0])
for u in range(1,(len(hash_sort)-1)):
	if hash_sort[u]!=hash_coll[v]:
		v=v+1		
		hash_coll.append(hash_sort[u])
		sheet_sorthashtag.cell(row= v+1,column=1).value = hash_sort[u]



tweet_hashtag_index=[]
for s in range (2,509):
	#sheet_sorthashtag.cell(row=s, column=2).value=0
	#sheet_sorthashtag.cell(row=s, column=3).value=0			
	for t in range(2,1256):
		if (sheet_new.cell(row=t, column=3).value.find(sheet_sorthashtag.cell(row=s, column=1).value) != -1):
	#fav_tweet & retweet_count aufsummieren


			#sheet_sorthashtag.cell(row=s, column=3).value = sheet_sorthashtag.cell(row=s, column=3).value + sheet_new.cell(row=t, column=6).value
			#sheet_sorthashtag.cell(row=s, column=4).value = sheet_sorthashtag.cell(row=s, column=4).value + sheet_new.cell(row=t, column=5).value
			tweet_hashtag_index.append('$'+str(t-2)+'$')
	string_t_h_index= str(tweet_hashtag_index)	
	#string_t_h_index = string_t_h_index[1:len(string_t_h_index)-1] =>	brackets rausnehemen
	sheet_sorthashtag.cell(row=s, column=2).value = string_t_h_index
	tweet_hashtag_index = []
savedElection.save("Election.xlsx")
newHashtag.save("Hashtag_list.xlsx")
	



xlsx2csv("Election.xlsx", open("Election.csv", "w+"))
xlsx2csv("Hashtag_list.xlsx", open("Hashtag_list.csv", "w+"))


